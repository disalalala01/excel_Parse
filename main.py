from openpyxl import load_workbook
import requests
import redis
import logging
import json
import time


cache = redis.Redis(host='localhost', port=6379, db=7)


def main():
    try:
        filename = '70000.xlsx'
        wb = load_workbook(filename)
        sheet_name = wb.sheetnames[0]
        sheet = wb.get_sheet_by_name(sheet_name)
        row = sheet.max_row
        last_row = f'B{row}'
        count = 0
        for cellObj in sheet['A2': last_row]:
            identifier = cellObj[0].value
            data = {}
            count += 1
            logging.info(f'COUNT : {count}')
            logging.info(f'IDENTIFIER: {identifier}')
            key = f'vig_{identifier}'

            if cache.get(key):
                logging.info(f'EXIST IDENTIFIER : {identifier}')
            else:
                url = "https://stat.gov.kz/api/juridical/counter/api/?bin={}&lang=ru".format(identifier)
                try:
                    r = requests.get(url)
                except requests.exceptions.SSLError:
                    r = requests.get(url)
                time.sleep(0.5)
                if r.status_code == 200:
                    resp = r.json()
                    if resp['success']:
                        logging.info('SUCCESS TRUE')
                        data['identifier'] = identifier
                        data['name'] = resp['obj']['name']
                        data['status'] = 'YES'
                        cache.set(key, json.dumps(data))
                    else:
                        logging.info('SUCCESS: FALSE')
                        data['identifier'] = identifier
                        data['name'] = None
                        data['status'] = 'NO'
                        cache.set(key, json.dumps(data))
                else:
                    data['identifier'] = identifier
                    data['name'] = None
                    data['status'] = None
                    cache.set(key, json.dumps(data))
            logging.info('-' * 50)
    except Exception as e:
        logging.exception(e)


if __name__ == '__main__':
    logging.basicConfig(handlers=[logging.FileHandler("vig.log", 'a+', 'utf-8')],
                        format='%(asctime)s %(levelname)s:%(message)s', level=logging.INFO)
    main()
